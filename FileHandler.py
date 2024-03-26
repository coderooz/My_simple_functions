import json, csv, os
from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfReader
from functions import DataHandlers as dh


def getFiles(file_path:str='', catg:int=0, extention=None, full_path:bool=False):
    """

    Args:
        file_path (str): Takes the path of the folder that needs to be looked into.
        catg (int): This parameter takes either 1 or 2, where 1 means to only look for files whereas 2 means to lik for directories. Default is 0. Which means both.
            
        
    Returns:
        (dict|list): Returns a list of files or directories in the given file.
    """
    if file_path=='':
        file_path = os.path.dirname(os.path.abspath(__file__))

    files = os.listdir(file_path)

    if full_path or catg==1 or extention!=None:
        files = [os.path.join(file_path, file) for file in files]
                    
    if catg == 1 or extention != None:
        if extention!=None:
            files = [i for i in files if (getExtention(i)) in extention]
        else: 
            files = [i for i in files if os.path.isfile(i)]

    elif catg == 2:
        files = [i for i in files if os.path.isdir(i)]
    elif catg == 0:
        pass
    else: raise ValueError('The parameter passed in catg is invalid. It only accepts 1 or 2 as valid parameter.')
    return files if full_path else [get_only_filename(i) for i in files]

def get_only_filename(full_path):
    """_summary_

    Args:
        full_path (str): _description_

    Returns:
        str: _description_
    """
    return os.path.basename(full_path)

def getFileCatg(file_path:str):
    """
    
    """        
    if os.path.isdir(file_path):
        return 2
    elif os.path.isfile(file_path):
        return 1
    return None

def splitFileName(file_path:str):
    """_summary_

    Args:
        file_path (str): _description_

    Returns:
        tuple: _description_
    """
    return os.path.splitext(get_only_filename(file_path))[0]

def getFileSize(file_path:str):
    """_summary_

    Args:
        file_path (str): _description_

    Returns:
        bytes: _description_
    """

    if os.path.isfile(file_path):
        return os.path.getsize(file_path)
    return False

def getExtention(file:str):
    """_summary_

    Args:
        file (str): _description_

    Returns:
        _type_: _description_
    """
    return os.path.splitext(file)[1].replace('.','')

def fileExists(file:str):
    """Checks if a file exists or not"""
    return True if os.path.exists(file) and os.path.isfile(file) else False

def write(file_name: str, data, separator='', emptyPervious:bool=False) -> None:
    """
    Appends data to a file.

    :param file_name: The name of the file.
    :param data: The data to be written to the file.
    :param separator: Optional separator to append after the data.
    """
    
    if emptyPervious: 
        open(file_name, 'w').close()

    if isinstance(data,(list, tuple, set)):
        [write(file_name, d, separator) for d in data]
    else:

        with open(file_name, 'a', encoding='utf-8', errors='ignore') as file:
            data_type = type(data)
            if data_type == str:
                file.write(data)
            elif data_type == dict:
                json.dump(data, file)
            else: TypeError('The data type provided is not supported.')
            file.write(separator)

def read(file_name: str, splitter=None, encode='utf-8', error='ignore', read_bytes:bool=False, decode_json=False):
    """
        Reads data from a file.

        :param file_name: The name of the file.
        :param splitter: Optional splitter to split the data.
        :param encode: Encoding of the file.
        :param error: How to handle encoding errors.
        :param decode_json: Whether to decode JSON-formatted data.

        :return: The read data.
    """
    mode = 'rb' if read_bytes else 'r'  
    with open(file_name, mode, encoding=encode, errors=error) as file:
        data = file.read()
        retdata = []
        if splitter is not None:
            for ret in dh.remove_empty_strings(data.split(splitter)):
                if decode_json:
                    ret = json.loads(ret)
                retdata.append(ret)
            return retdata
        else:
            if decode_json:
                return json.loads(data)
            else:
                return data

def write_csv(file_name: str, data, header=None):
    """
    Writes data to a CSV file.

    :param file_name: The name of the CSV file.
    :param data: The data to be written to the CSV file.
    :param header: Optional header for the CSV file.
    """
    with open(file_name, 'w', newline='', encoding='utf-8', errors='ignore') as csvfile:
        csv_writer = csv.writer(csvfile)
        if header:
            csv_writer.writerow(header)
        csv_writer.writerows(data)

def read_csv(file_name: str, header:bool=False):
    """
        read_csv()
        ----------
        Reads data from a CSV file.
        
        Parameter:
        - file_name (str): The name of the CSV file.
        - header (bool, optional): 

        :return: The read data.
    """
    with open(file_name, 'r', encoding='utf-8', errors='ignore') as csvfile:
        csv_reader = csv.reader(csvfile)
        data = [row for row in csv_reader]
        if header == True:
            return dh.dlist_dict(dh.list_dlist(data[1:], data[0]))
        return data
    
def write_excel(filename, data, sheetname='Sheet 1'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheetname
    if isinstance(data, list) and isinstance(data[0], dict):
        pass
    elif isinstance(data, list) and isinstance(data[0], list):
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                sheet.cell(row=i+1, column=j+1, value=value)
    elif isinstance(data, dict):
        pass
    workbook.save(filename)

def read_excel(file_name: str, sheet_name=None, organize:bool=False):
    """
    Reads data from an Excel (XLSX) file.

    :param file_name: The name of the Excel file.
    :param sheet_name: Optional sheet name for the Excel file.

    :return: The read data.
    """
    workbook = load_workbook(file_name)
    sheet_data = {}
    if sheet_name!=None and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = [list(row) for row in sheet.iter_rows(values_only=True)]
        if organize:
            sheet_data = dh.dlist_dict(dh.list_dlist(sheet_data[1:], sheet_data[0]))
    else:    
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = [list(row) for row in sheet.iter_rows(values_only=True)]
            sheet_data[sheet_name] = dh.dlist_dict(dh.list_dlist(data[1:], data[0])) if organize else data

    return sheet_data

def read_pdf(file_name: str):
    """
    Reads text data from a PDF file.

    :param file_name: The name of the PDF file.

    :return: The read text data.
    """
    with open(file_name, 'rb') as pdf_file:
        pdf_reader = PdfReader(pdf_file)
        text = ''
        for page in pdf_reader.pages:
            text += page.extract_text()
        return text

def read_har(fileName:str):
    entries = []
    urls = []
    data = dh.decode_json(read(fileName).replace('\n','').replace('\r', '').replace('\t', ''))['log']
    for i in data['entries'][1:]:
        full_url = i['request']['url']
        if full_url not in urls and i['request']['headers'][0]['value']=='www.nseindia.com' and '/api' in full_url and full_url != "https://www.nseindia.com/api/marketStatus":
            headers = {hr['name']:hr['value'] for hr in i['request']['headers']}
            entries.append({'file':fileName,'full_url':full_url, 'headers':headers, 'response': i['response']['content']})
            urls.append(full_url)
    return urls, entries
        

